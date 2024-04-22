##########################################################################################################################################################################################################################################
#    Script Description:
#        This script opens a list of network message data files (vsb or db2) and makes a spreadsheet with summary information about each file listing the first and last timestamp
#         as well as a list of all the record count for each ArbID on each network
#
###########################################################################################################################################################################################################################################
#    Script Inputs: When you run this script you will be prompted with 2 file open dialog windows. The first asks for a config file with extension *.asl; the second asks
#    for a list of one or more network data files (*.vsb, *.db2)
#
#         Script config file GenerateFileInfoSummary_Config.asl is a JSON file used to configure the script. This file has the following keys:
#
#             TemplateFilename - Excel file that is used as a template for generating the report. You can change the colors or bolding of the template
#								 and your changes will be used for future reports. 
#
#             DeleteTempDB2FileAfterExecution - Flag can be TRUE or FALSE. If TRUE it will delete any temporary db2 if they were made by the script. Any db2 files that were
#                                               input to the script will not be deleted
#
#        Network data file(s) list. Files can be in vsb or db2 format or mixture of these.
#
##########################################################################################################################################################################################################################################
#    Script Outputs:
#        The script outputs a *.xlsx file based on the TemplateFilename but with summary information about the vsb or db2 files analyzed. 
#        
##########################################################################################################################################################################################################################################

import numpy as np
import sys
import os
import logging
import json
import sys, os

from openpyxl.reader.excel import load_workbook
from UtilityFunctions import ConvertNetworkStringToID
from MsgFileClass import msgFiles
from ICS_VSBIO import ICSFileInterfaceLibrary
from shutil import copyfile

from datetime import datetime, timezone

slFilePath = ICSFileInterfaceLibrary.get_config_file()
inputFilePaths = ICSFileInterfaceLibrary.get_input_file_list()

ReportGenTimeStamp = datetime.now().strftime("%m-%d-%y_%H-%M-%S")

logging.basicConfig(level=logging.INFO)
log = logging.getLogger(__name__)
loggingPath = "IPA.log"
print(slFilePath)
is_wivi35 = ICSFileInterfaceLibrary.is_running_on_wivi_server() and os.path.splitext(sys.argv[1])[1].lower() == '.json'

if is_wivi35:	
	config = json.loads(slFilePath)
	ipaInstanceConfig = json.load(open(sys.argv[1]))
	config['output_dir'] = ipaInstanceConfig['output_dir']
	loggingPath = os.path.join(ipaInstanceConfig["output_dir"], "IPA.log")
	OutputFilePath = config['output_dir']
	TemplateFilenameAndPath = os.path.join(OutputFilePath, config["TemplateFilename"])
	OutputFilename = "vsbFileInfoSummary_" + str(ReportGenTimeStamp) + ".xlsx"
	OutputFilenameAndPath = os.path.join(OutputFilePath ,OutputFilename)
else:
	configFile = open(slFilePath)
	config = json.load(configFile)
	config['output_dir'] = os.getcwd()
	OutputFilePath = config['output_dir']
	TemplateFilenameAndPath = os.path.join(OutputFilePath, config["TemplateFilename"])
	OutputFilename = "vsbFileInfoSummary_" + str(ReportGenTimeStamp) + ".xlsx"
	OutputFilenameAndPath = os.path.join(OutputFilePath ,OutputFilename)

handler = logging.FileHandler(loggingPath)
handler.setLevel(logging.INFO)
# create a logging format
formatter = logging.Formatter('%(asctime)s - %(name)s - %(message)s')
handler.setFormatter(formatter)
log.addHandler(handler)
log.info("Hello")

log.info(slFilePath)

# now go through each file, look for hits and log hits to dsr file
# create a list for vsbFiles

log.info("Analyzing input files")
input_msg_Files = msgFiles(inputFilePaths, ReportGenTimeStamp, OutputFilePath, log)

#now write the vsbFiles info to the template Excel Spreadsheet
log.info("Creating Excel output file")
wb = load_workbook(TemplateFilenameAndPath)	
DataEntryRow = 3

if(len(input_msg_Files.FilesListSorted) > 0):	
	if not('vsbStatSummary' in wb.sheetnames):
		log.info(OutputFilename + " doesn't have a sheet named vsbStatSummary")
		wb.save(OutputFilenameAndPath)
		wb.close()
		quit()
	ws = wb.get_sheet_by_name('vsbStatSummary')

	#Read in the header rows 
	ColumnHeadersFromTemplateSpreadsheet = []
	i = 0
	while (ws.cell(row=2, column=1 + i).value != None):
		ColumnHeadersFromTemplateSpreadsheet.append(str(ws.cell(row=2, column=1 + i).value))
		i = i + 1
	NumberOfDIDColumnsInLog = i-1

	log.info("Begin writing report data to DTCReport sheet")

	for a in range(len(input_msg_Files.FilesListSorted)):
		log.info("Writing network info for File number " + str(a))
		for b in range(len(input_msg_Files.FilesListSorted[a].FileNetworks)):
			for c in range(len(input_msg_Files.FilesListSorted[a].FileNetworks[b].network_msg_ids_hex)):
				ws.cell(row=DataEntryRow, column=1, value=input_msg_Files.FilesListSorted[a].FileStartTime)
				ws.cell(row=DataEntryRow, column=2, value=input_msg_Files.FilesListSorted[a].FileEndTime)
				ws.cell(row=DataEntryRow, column=3, value=input_msg_Files.FilesListSorted[a].InputFileName)
				ws.cell(row=DataEntryRow, column=4, value=input_msg_Files.FilesListSorted[a].FileNetworks[b].network_name)
				ws.cell(row=DataEntryRow, column=5, value=input_msg_Files.FilesListSorted[a].FileNetworks[b].network_msg_ids_hex[c])
				ws.cell(row=DataEntryRow, column=6, value=input_msg_Files.FilesListSorted[a].FileNetworks[b].network_msg_num_msgs[c])
				ws.cell(row=DataEntryRow, column=7, value=input_msg_Files.FilesListSorted[a].FileNetworks[b].network_msg_id_min_periods[c])
				ws.cell(row=DataEntryRow, column=8, value=input_msg_Files.FilesListSorted[a].FileNetworks[b].network_msg_id_max_periods[c])
				DataEntryRow = DataEntryRow + 1

		if ( os.path.isfile(input_msg_Files.FilesListSorted[a].DB_FileName) and (input_msg_Files.FilesListSorted[a].FileCreatedByClass) \
		    and (config["DeleteTempDB2FileAfterExecution"] == "TRUE")  ):
			os.remove(input_msg_Files.FilesListSorted[a].DB_FileName)

	log.info("Saving report file")
	if ICSFileInterfaceLibrary.is_running_on_wivi_server():
		wb.save(OutputFilename)
	else:
		wb.save(OutputFilenameAndPath)
		
	wb.close()
	log.info("Goodbye")

