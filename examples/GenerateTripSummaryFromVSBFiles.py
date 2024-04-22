##########################################################################################################################################################################################################################################
#    Script Description:
#        This script opens a list of network message data files (vsb or db2) and makes a spreadsheet with summary information about each file listing the first and last timestamp
#         as well as a list of all the record count for each ArbID on each network
#
###########################################################################################################################################################################################################################################
#    Script Inputs: When you run this script you will be prompted with 2 file open dialog windows. The first asks for a config file with extension *.asl; the second asks
#    for a list of one or more network data files (*.vsb, *.db2)
#
#         Script config file GenerateFileInfoSummary_Config.asl is a JSON file used to configure the script. This file has the following keys:
#
#             TemplateFilename - Excel file that is used as a template for generating the report. You can change the colors or bolding of the template
#								 and your changes will be used for future reports. 
#
#             DeleteTempDB2FileAfterExecution - Flag can be TRUE or FALSE. If TRUE it will delete any temporary db2 if they were made by the script. Any db2 files that were
#                                               input to the script will not be deleted
#
#        Network data file(s) list. Files can be in vsb or db2 format or mixture of these.
#
##########################################################################################################################################################################################################################################
#    Script Outputs:
#        The script outputs a *.xlsx file based on the TemplateFilename but with summary information about the vsb or db2 files analyzed. 
#        
##########################################################################################################################################################################################################################################

import numpy as np
import sys
import os   
import logging
import json
import sys, os

from openpyxl.reader.excel import load_workbook
from UtilityFunctions import ConvertNetworkStringToID
from MsgFileClass import msgFiles
from ICS_VSBIO import ICSFileInterfaceLibrary
from shutil import copyfile
from zipfile import ZipFile
from datetime import datetime, timezone
from TripsSummaryClass import tripsSummary

slFilePath = ICSFileInterfaceLibrary.get_config_file()
inputFilePaths = ICSFileInterfaceLibrary.get_input_file_list()
ReportGenTimeStamp = datetime.now().strftime("%m-%d-%y_%H-%M-%S")

is_wivi35 = ICSFileInterfaceLibrary.is_running_on_wivi_server() and os.path.splitext(sys.argv[1])[1].lower() == '.json'

# for wivi prior to 3.5, *.asl config file does not include path, it gets the path from the path of __main__.py
if is_wivi35:
	config = json.loads(slFilePath)
	ipaInstanceConfig = json.load(open(sys.argv[1]))
	config['output_dir'] = ipaInstanceConfig['output_dir']
	loggingPath = os.path.join(ipaInstanceConfig["output_dir"], "IPA.log")
	TemplateFileName = config["TemplateFilename"]
	OutputFilePath = config['output_dir']
	OutputFilename = "vsbFileInfoSummary_" + str(ReportGenTimeStamp) + ".xlsx"
	OutputFilenameAndPath = os.path.join(OutputFilePath ,OutputFilename)
	TemplateFileNameAndPath = os.path.join(config['output_dir'], TemplateFileName)
	ScriptZipFileNameAndPath = 	sys.argv[0]
	StatsXLSXFilenameAndPath  = os.path.join(config['output_dir'], "./SigStats_" + ReportGenTimeStamp + ".xlsx")
	with ZipFile(ScriptZipFileNameAndPath, 'r') as zipObject:
		listOfFileNames = zipObject.namelist()
		for fileName in listOfFileNames:
			if fileName == TemplateFileName:
				# Extract a single file from zip
				zipObject.extract(fileName, config['output_dir'])			
else:
	configFile = open(slFilePath)
	config = json.load(configFile)
	config['output_dir'] = os.getcwd()
	loggingPath = "IPA.log"
	TemplateFileName = config["TemplateFilename"]
	OutputFilePath = config['output_dir']
	TemplateFileNameAndPath = os.path.join(OutputFilePath, config["TemplateFilename"])
	OutputFilename = "vsbFileInfoSummary_" + str(ReportGenTimeStamp) + ".xlsx"
	OutputFilenameAndPath = os.path.join(OutputFilePath ,OutputFilename)

logging.basicConfig(level=logging.INFO)
log = logging.getLogger(__name__)
print(slFilePath)
handler = logging.FileHandler(loggingPath)
handler.setLevel(logging.INFO)
# create a logging format
formatter = logging.Formatter('%(asctime)s - %(name)s - %(message)s')
handler.setFormatter(formatter)
log.addHandler(handler)
log.info("Hello")
log.info("TemplateFileNameAndPath: " + TemplateFileNameAndPath)
log.info("OutputFileNameAndPath: " + OutputFilenameAndPath)

# now go through each file, look for hits and log hits to dsr file
# create a list for vsbFiles
wb = load_workbook(TemplateFileNameAndPath)	

log.info("Analyzing input files")
input_msg_Files = msgFiles(inputFilePaths, ReportGenTimeStamp, OutputFilePath, log)

trips_summary = tripsSummary(input_msg_Files.FilesListSorted, ReportGenTimeStamp, log)

#now write the vsbFiles info to the template Excel Spreadsheet
log.info("Creating Excel output file")

if(len(input_msg_Files.FilesListSorted) > 0):	
	if config['GenerateTripSummary'] == "TRUE":
		DataEntryRow = 3
		if not('TripSummary' in wb.sheetnames):
			log.info(OutputFilename + " doesn't have a sheet named TripSummary")
			wb.save(OutputFilenameAndPath)
			wb.close()
			quit()
		ws = wb.get_sheet_by_name('TripSummary')

		#Read in the header rows 
		ColumnHeadersFromTemplateSpreadsheet = []
		i = 0
		while (ws.cell(row=2, column=2 + i).value != None):
			ColumnHeadersFromTemplateSpreadsheet.append(str(ws.cell(row=2, column=1 + i).value))
			i = i + 1

		log.info("Begin writing report data to TripSummary sheet")

		for a in range(trips_summary.NumberOfTrips):
			log.info("Writing network info for Trip number " + str(a))
			ws.cell(row=DataEntryRow, column=2, value= a)
			ws.cell(row=DataEntryRow, column=3, value=trips_summary.TripStartTime[a])
			ws.cell(row=DataEntryRow, column=4, value=trips_summary.TripEndTime[a])
			ws.cell(row=DataEntryRow, column=5, value=trips_summary.TripDurationMinutes[a])
			ws.cell(row=DataEntryRow, column=6, value=trips_summary.FirstMsgFileNameInTrip[a])
			ws.cell(row=DataEntryRow, column=7, value=trips_summary.LastMsgFileNameInTrip[a])
			ws.cell(row=DataEntryRow, column=8, value=trips_summary.NumberOfFilesInTrip[a])
			ws.cell(row=DataEntryRow, column=9, value=trips_summary.MinTimeGapBetweenFilesInTrip[a])
			ws.cell(row=DataEntryRow, column=10, value=trips_summary.MaxTimeGapBetweenFilesInTrip[a])
			DataEntryRow = DataEntryRow + 1

	if config['GenerateMsgFileSummary'] == "TRUE":
		DataEntryRow = 3
		if not('MsgFileSummary' in wb.sheetnames):
			log.info(OutputFilename + " doesn't have a sheet named MsgFileSummary")
			wb.save(OutputFilenameAndPath)
			wb.close()
			quit()
		ws = wb.get_sheet_by_name('MsgFileSummary')

		#Read in the header rows 
		ColumnHeadersFromTemplateSpreadsheet = []
		i = 0
		while (ws.cell(row=2, column=1 + i).value != None):
			ColumnHeadersFromTemplateSpreadsheet.append(str(ws.cell(row=2, column=1 + i).value))
			i = i + 1

		log.info("Begin writing report data to MsgFileSummary sheet")

		for a in range(len(input_msg_Files.FilesListSorted)):
			log.info("Writing message file info for File number " + str(a))
			for b in range(len(input_msg_Files.FilesListSorted[a].FileNetworks)):
				ws.cell(row=DataEntryRow, column=2, value=a)
				ws.cell(row=DataEntryRow, column=3, value=input_msg_Files.FilesListSorted[a].FileStartTime)
				ws.cell(row=DataEntryRow, column=4, value=input_msg_Files.FilesListSorted[a].FileEndTime)
				ws.cell(row=DataEntryRow, column=5, value=input_msg_Files.FilesListSorted[a].FileDurationInMin)					
				ws.cell(row=DataEntryRow, column=6, value=input_msg_Files.FilesListSorted[a].InputFileName)
				ws.cell(row=DataEntryRow, column=7, value=input_msg_Files.FilesListSorted[a].FileNetworks[b].network_name)
				ws.cell(row=DataEntryRow, column=8, value=input_msg_Files.FilesListSorted[a].FileNetworks[b].network_tot_number_messages_in_file)
				ws.cell(row=DataEntryRow, column=9, value=input_msg_Files.FilesListSorted[a].FileNetworks[b].numberOfArbIDs)
				log.info("File duration in minutes: " + str(input_msg_Files.FilesListSorted[a].FileDurationInMin))
				if input_msg_Files.FilesListSorted[a].FileDurationInMin > 0:
					ws.cell(row=DataEntryRow, column=10, value=(input_msg_Files.FilesListSorted[a].FileNetworks[b].network_tot_number_messages_in_file / (input_msg_Files.FilesListSorted[a].FileDurationInMin*60)))
				else:
					ws.cell(row=DataEntryRow, column=10, value="N/A")
				ws.cell(row=DataEntryRow, column=11, value=input_msg_Files.FilesListSorted[a].FileNetworks[b].network_data_start_time)
				ws.cell(row=DataEntryRow, column=12, value=input_msg_Files.FilesListSorted[a].FileNetworks[b].network_data_end_time)
				ws.cell(row=DataEntryRow, column=13, value=input_msg_Files.FilesListSorted[a].time_gap_from_prev_file)
				ws.cell(row=DataEntryRow, column=14, value=input_msg_Files.FilesListSorted[a].FileNetworks[b].network_min_time_gap)
				ws.cell(row=DataEntryRow, column=15, value=input_msg_Files.FilesListSorted[a].FileNetworks[b].network_max_time_gap)
				DataEntryRow = DataEntryRow + 1

	if config['GenerateArbIdSummaryPerNetwork'] == "TRUE":
		DataEntryRow = 3
		if not('ArbIDSummary' in wb.sheetnames):
			log.info(OutputFilename + " doesn't have a sheet named ArbIDSummary")
			wb.save(OutputFilenameAndPath)
			wb.close()
			quit()
		ws = wb.get_sheet_by_name('ArbIDSummary')

		#Read in the header rows 
		ColumnHeadersFromTemplateSpreadsheet = []
		i = 0
		while (ws.cell(row=2, column=1 + i).value != None):
			ColumnHeadersFromTemplateSpreadsheet.append(str(ws.cell(row=2, column=1 + i).value))
			i = i + 1
		NumberOfDIDColumnsInLog = i-1

		log.info("Begin writing report data to DTCReport sheet")

		for a in range(len(input_msg_Files.FilesListSorted)):
			log.info("Writing network info for File number " + str(a))
			for b in range(len(input_msg_Files.FilesListSorted[a].FileNetworks)):
				for c in range(len(input_msg_Files.FilesListSorted[a].FileNetworks[b].network_msg_ids_hex)):
					ws.cell(row=DataEntryRow, column=2, value=input_msg_Files.FilesListSorted[a].FileStartTime)
					ws.cell(row=DataEntryRow, column=3, value=input_msg_Files.FilesListSorted[a].FileEndTime)
					ws.cell(row=DataEntryRow, column=4, value=input_msg_Files.FilesListSorted[a].InputFileName)
					ws.cell(row=DataEntryRow, column=5, value=input_msg_Files.FilesListSorted[a].FileNetworks[b].network_name)
					ws.cell(row=DataEntryRow, column=6, value=input_msg_Files.FilesListSorted[a].FileNetworks[b].network_msg_ids_hex[c])
					ws.cell(row=DataEntryRow, column=7, value=input_msg_Files.FilesListSorted[a].FileNetworks[b].network_msg_num_msgs[c])
					ws.cell(row=DataEntryRow, column=8, value=input_msg_Files.FilesListSorted[a].FileNetworks[b].network_msg_id_min_periods[c])
					ws.cell(row=DataEntryRow, column=9, value=input_msg_Files.FilesListSorted[a].FileNetworks[b].network_msg_id_max_periods[c])
					DataEntryRow = DataEntryRow + 1

	log.info("Saving report file")
	wb.save(OutputFilenameAndPath)		
	wb.close()

	for msg_File in input_msg_Files.FilesListSorted:
		if ( os.path.isfile(msg_File.DB_FileName) and (config["DeleteTempDB2FileAfterExecution"] == "TRUE")  ):
			os.remove(msg_File.DB_FileName)
			log.info("Removed file: " + msg_File.DB_FileName )

	if is_wivi35 and os.path.isfile(TemplateFileNameAndPath):
		os.remove(TemplateFileNameAndPath)
	log.info("Goodbye")

